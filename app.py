#!/usr/bin/env python3
"""
BaseLinker Sales Dashboard - FastAPI Application
Displays all sold products with real-time updates, Excel export, and activity tracking
Enhanced: category detection, purchase orders, revenue tracking, sales channels, time frame filtering
Authentication: PIN/email login, admin panel, user sessions, activity logging
"""

import os
import re
import json
import time
import asyncio
import threading
import requests
import uuid
import secrets
import jwt
import bcrypt
import logging
import unicodedata
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from collections import defaultdict
from enum import Enum

logger = logging.getLogger("sales-dashboard")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def strip_diacritics(text: str) -> str:
    """Remove diacritics/accents for accent-insensitive search."""
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))

# User is in Poland — all date filtering must use Polish timezone
POLISH_TZ = ZoneInfo("Europe/Warsaw")
from io import BytesIO
from typing import Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, Response, BackgroundTasks, Query, Depends, HTTPException, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel

# Configuration from environment variables
BASELINKER_API_KEY = os.getenv("BASELINKER_API_KEY", "")
BASELINKER_INVENTORY_ID = int(os.getenv("BASELINKER_INVENTORY_ID", "58952"))
BASELINKER_API_URL = "https://api.baselinker.com/connector.php"
WYSLANE_STATUS_ID = 273568   # Wysłane (Shipped)
SPAKOWANE_STATUS_ID = 273928  # Spakowane (Packed)
ANULOWANE_STATUS_ID = 273569  # Anulowane (Cancelled)
FINANCIAL_STATUS_IDS = {WYSLANE_STATUS_ID, SPAKOWANE_STATUS_ID}  # Statuses that count toward financials
EXCLUDED_STATUS_IDS = {ANULOWANE_STATUS_ID}  # Statuses that should NEVER count
DATABASE_URL = os.getenv("DATABASE_URL", "")

# Google Sheets cost loading
GOOGLE_CREDENTIALS_JSON = os.getenv("GOOGLE_CREDENTIALS_JSON", "")
IMPORT_SHEET_ID = os.getenv("IMPORT_SHEET_ID", "1parHGahhvO6qnAnsvu4MNq18yl3M53Dzs6GCjry1yMg")
FALLBACK_MARKUP_GROSS = 3.444  # cost = gross_price / 3.444 (2.8x net markup + 23% VAT)

# JWT Configuration
JWT_SECRET = os.getenv("JWT_SECRET")
if not JWT_SECRET:
    import sys
    print("FATAL: JWT_SECRET environment variable is not set. Exiting.", file=sys.stderr)
    sys.exit(1)
JWT_ALGORITHM = "HS256"
JWT_EXPIRY_HOURS = 72
security = HTTPBearer(auto_error=False)

# Currency conversion — BaseLinker returns price_brutto in the ORDER's currency
EXCHANGE_RATES_AS_OF = "2026-03-01"

# VAT rates by currency — used for net revenue estimation
# PLN orders are Polish (23%), others use origin country standard rate
VAT_RATES_BY_CURRENCY = {
    'PLN': 0.23,
    'EUR': 0.23,   # Default EU (most EU sales are Polish-VAT registered)
    'CZK': 0.21,
    'HUF': 0.27,
    'SEK': 0.25,
    'DKK': 0.25,
    'NOK': 0.25,
    'RON': 0.19,
    'BGN': 0.20,
    'GBP': 0.20,
    'USD': 0.0,    # Non-EU — no VAT
    'CHF': 0.0,    # Non-EU
}
DEFAULT_VAT_RATE = 0.23
EXCHANGE_RATES_TO_PLN = {
    'PLN': 1.0,
    'EUR': 4.18,
    'USD': 3.95,
    'GBP': 5.00,
    'CZK': 0.17,
    'HUF': 0.0105,
    'SEK': 0.37,
    'DKK': 0.56,
    'NOK': 0.36,
    'CHF': 4.45,
    'RON': 0.84,
    'BGN': 2.14,
    'HRK': 0.55,
    'RUB': 0.043,
    'UAH': 0.095,
}


def convert_to_pln(amount: float, currency: str) -> float:
    """Convert amount from order currency to PLN."""
    if not currency or currency.upper() == 'PLN':
        return amount
    rate = EXCHANGE_RATES_TO_PLN.get(currency.upper())
    if rate:
        return amount * rate
    logger.warning(f"Unknown currency '{currency}' — treating as PLN")
    return amount

# Data cache
cache = {
    "data": None,
    "last_updated": None,
    "next_refresh": None,
    "is_refreshing": False,
    "purchase_orders": [],
    "raw_orders": None,
    "inventory": None,
    "po_items_by_bl_id": None,
    "order_sources": {},
    "all_recent_orders": None,
    "costs": ({}, {}),
    "full_inventory": [],  # All products/variants for search (name, sku, stock, image, bl_id)
}

REFRESH_INTERVAL = 1800  # 30 minutes in seconds

# Database connection pool
db_pool = None


def init_db_pool():
    """Initialize the ThreadedConnectionPool."""
    global db_pool
    if not DATABASE_URL:
        return
    try:
        from psycopg2.pool import ThreadedConnectionPool
        db_pool = ThreadedConnectionPool(2, 10, DATABASE_URL)
        logger.info("Database connection pool initialized (min=2, max=10)")
    except Exception as e:
        logger.error(f"Database pool initialization error: {e}")
        db_pool = None


def get_db_connection():
    """Get a connection from the pool. Caller MUST return it via put_db_connection()."""
    if db_pool is None:
        return None
    try:
        conn = db_pool.getconn()
        conn.autocommit = False
        return conn
    except Exception as e:
        logger.error(f"Database connection error: {e}")
        return None


def put_db_connection(conn):
    """Return a connection to the pool."""
    if conn is not None and db_pool is not None:
        try:
            db_pool.putconn(conn)
        except Exception:
            pass


def init_database():
    """Initialize database tables"""
    conn = get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS activity_log (
                id SERIAL PRIMARY KEY,
                timestamp TIMESTAMPTZ DEFAULT NOW(),
                action VARCHAR(50),
                total_orders INT,
                total_variants INT,
                total_units_sold INT,
                details JSONB
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sales_snapshot (
                id SERIAL PRIMARY KEY,
                timestamp TIMESTAMPTZ DEFAULT NOW(),
                sku VARCHAR(100),
                product_name TEXT,
                units_sold INT,
                current_stock INT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                email VARCHAR(255) UNIQUE NOT NULL,
                display_name VARCHAR(255),
                pin_hash VARCHAR(255),
                password_hash VARCHAR(255),
                role VARCHAR(20) DEFAULT 'user',
                is_banned BOOLEAN DEFAULT FALSE,
                must_change_password BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                updated_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_sessions (
                id SERIAL PRIMARY KEY,
                user_id INT REFERENCES users(id) ON DELETE CASCADE,
                token_jti VARCHAR(255) UNIQUE NOT NULL,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                last_heartbeat TIMESTAMPTZ DEFAULT NOW(),
                expires_at TIMESTAMPTZ NOT NULL,
                ip_address VARCHAR(45),
                user_agent TEXT,
                is_active BOOLEAN DEFAULT TRUE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_activity (
                id SERIAL PRIMARY KEY,
                user_id INT REFERENCES users(id) ON DELETE CASCADE,
                session_jti VARCHAR(255),
                timestamp TIMESTAMPTZ DEFAULT NOW(),
                action VARCHAR(100),
                details JSONB
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS password_reset_requests (
                id SERIAL PRIMARY KEY,
                user_id INT REFERENCES users(id) ON DELETE CASCADE,
                requested_at TIMESTAMPTZ DEFAULT NOW(),
                reason TEXT,
                status VARCHAR(20) DEFAULT 'pending',
                fulfilled_at TIMESTAMPTZ,
                fulfilled_by INT,
                temp_password VARCHAR(255)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS pin_attempts (
                id SERIAL PRIMARY KEY,
                ip_address VARCHAR(45),
                attempted_at TIMESTAMPTZ DEFAULT NOW(),
                success BOOLEAN DEFAULT FALSE
            )
        """)
        # Permanent order storage — survives BaseLinker's 3-month auto-archive
        cur.execute("""
            CREATE TABLE IF NOT EXISTS bl_orders (
                order_id BIGINT PRIMARY KEY,
                date_add BIGINT,
                date_confirmed BIGINT,
                status_id INTEGER,
                order_source VARCHAR(50),
                order_source_id INTEGER,
                currency VARCHAR(10),
                delivery_price NUMERIC(12,2),
                order_data JSONB NOT NULL,
                first_seen_at TIMESTAMPTZ DEFAULT NOW(),
                updated_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_bl_orders_date_add ON bl_orders(date_add)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_bl_orders_status ON bl_orders(status_id)")
        # Backfill status_id from order_data JSONB for rows where it's NULL
        cur.execute("""
            UPDATE bl_orders SET status_id = (order_data->>'order_status_id')::INTEGER
            WHERE status_id IS NULL AND order_data->>'order_status_id' IS NOT NULL
        """)
        conn.commit()

        # Seed admin user
        admin_pin = "4523"
        pin_hash = bcrypt.hashpw(admin_pin.encode(), bcrypt.gensalt()).decode()
        cur.execute("""
            INSERT INTO users (email, display_name, pin_hash, role, must_change_password)
            VALUES (%s, %s, %s, %s, TRUE)
            ON CONFLICT (email) DO NOTHING
        """, ('glamova.hdht@gmail.com', 'Admin', pin_hash, 'admin'))
        conn.commit()
        cur.close()
        logger.info("Database tables initialized")
    except Exception as e:
        logger.error(f"Database init error: {e}")
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


def save_orders_to_db(orders: list):
    """Save/update orders in the database so they survive BaseLinker's 3-month archive."""
    conn = get_db_connection()
    if not conn or not orders:
        return 0
    try:
        cur = conn.cursor()
        saved = 0
        for order in orders:
            oid = order.get('order_id')
            if not oid:
                continue
            cur.execute("""
                INSERT INTO bl_orders (order_id, date_add, date_confirmed, status_id,
                    order_source, order_source_id, currency, delivery_price, order_data)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (order_id) DO UPDATE SET
                    status_id = EXCLUDED.status_id,
                    order_data = EXCLUDED.order_data,
                    updated_at = NOW()
            """, (
                oid,
                order.get('date_add', 0),
                order.get('date_confirmed', 0),
                order.get('order_status_id') if order.get('order_status_id') is not None else order.get('status_id'),
                order.get('order_source', ''),
                order.get('order_source_id', 0),
                (order.get('currency', '') or 'PLN').upper(),
                float(order.get('delivery_price', 0) or 0),
                json.dumps(order)
            ))
            saved += 1
        conn.commit()
        cur.close()
        logger.info(f"Saved {saved} orders to database")
        return saved
    except Exception as e:
        logger.error(f"Error saving orders to DB: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        return 0
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


def load_orders_from_db(min_date_add: int = 0) -> list:
    """Load ALL stored orders from the database (including ones BaseLinker has archived).
    Returns list of order dicts (same format as BaseLinker API response).
    """
    conn = get_db_connection()
    if not conn:
        return []
    try:
        cur = conn.cursor('orders_server_cursor')
        cur.execute("SELECT order_data FROM bl_orders WHERE date_add >= %s ORDER BY date_add DESC", (min_date_add,))
        orders = []
        while True:
            batch = cur.fetchmany(500)
            if not batch:
                break
            for row in batch:
                orders.append(json.loads(row[0]) if isinstance(row[0], str) else row[0])
        cur.close()
        logger.info(f"Loaded {len(orders)} orders from database")
        return orders
    except Exception as e:
        logger.error(f"Error loading orders from DB: {e}")
        return []
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


def get_db_order_count() -> int:
    """Get count of orders stored in database."""
    conn = get_db_connection()
    if not conn:
        return 0
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM bl_orders")
        count = cur.fetchone()[0]
        cur.close()
        return count
    except Exception:
        return 0
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


def log_activity(action: str, data: dict = None):
    """Log activity to database"""
    conn = get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO activity_log (action, total_orders, total_variants, total_units_sold, details)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            action,
            data.get('total_orders', 0) if data else 0,
            data.get('total_variants', 0) if data else 0,
            data.get('total_units_sold', 0) if data else 0,
            json.dumps(data) if data else None
        ))
        conn.commit()
        cur.close()
    except Exception as e:
        logger.error(f"Log activity error: {e}")
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


def save_sales_snapshot(products: list):
    """Save current sales data snapshot"""
    conn = get_db_connection()
    if not conn or not products:
        return
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM sales_snapshot WHERE timestamp < NOW() - INTERVAL '90 days'")
        sorted_products = sorted(products, key=lambda x: x.get('units_sold', 0), reverse=True)
        for p in sorted_products[:50]:
            cur.execute("""
                INSERT INTO sales_snapshot (sku, product_name, units_sold, current_stock)
                VALUES (%s, %s, %s, %s)
            """, (p['sku'], p['product_name'][:200], p['units_sold'], p['current_stock']))
        conn.commit()
        cur.close()
    except Exception as e:
        logger.error(f"Save snapshot error: {e}")
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


# ========== AUTH HELPER FUNCTIONS ==========

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(password.encode(), hashed.encode())
    except Exception:
        return False

def verify_pin(pin: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pin.encode(), hashed.encode())
    except Exception:
        return False

def get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"

def create_session(user_id: int, request: Request) -> str:
    jti = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    expires = now + timedelta(hours=JWT_EXPIRY_HOURS)
    payload = {"sub": str(user_id), "jti": jti, "exp": expires, "iat": now}
    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

    conn = get_db_connection()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO user_sessions (user_id, token_jti, expires_at, ip_address, user_agent)
                VALUES (%s, %s, %s, %s, %s)
            """, (user_id, jti, expires, get_client_ip(request),
                  str(request.headers.get("user-agent", ""))[:500]))
            conn.commit()
            cur.close()
        except Exception as e:
            logger.error(f"Session creation error: {e}")
            try: conn.rollback()
            except Exception: pass
        finally:
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass
            put_db_connection(conn)
    return token

async def get_current_user(request: Request, credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = int(payload.get("sub", 0))
        jti = payload.get("jti")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")

    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT u.id, u.email, u.display_name, u.role, u.is_banned, u.must_change_password,
                   s.is_active
            FROM users u
            JOIN user_sessions s ON s.user_id = u.id AND s.token_jti = %s
            WHERE u.id = %s
        """, (jti, user_id))
        row = cur.fetchone()
        cur.close()

        if not row:
            raise HTTPException(status_code=401, detail="Session not found")
        if row[4]:  # is_banned
            raise HTTPException(status_code=403, detail="Account banned")
        if not row[6]:  # session not active
            raise HTTPException(status_code=401, detail="Session expired")

        return {
            "id": row[0], "email": row[1], "display_name": row[2],
            "role": row[3], "is_banned": row[4], "must_change_password": row[5],
            "jti": jti
        }
    except HTTPException:
        raise
    except Exception as e:
        try: conn.rollback()
        except Exception: pass
        raise HTTPException(status_code=500, detail=f"Auth error: {e}")
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

async def require_admin(user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

def log_user_activity(user_id: int, jti: str, action: str, details: dict = None):
    conn = get_db_connection()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO user_activity (user_id, session_jti, action, details)
            VALUES (%s, %s, %s, %s)
        """, (user_id, jti, action, json.dumps(details) if details else None))
        conn.commit()
        cur.close()
    except Exception as e:
        logger.error(f"Activity log error: {e}")
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


# ========== PYDANTIC MODELS ==========

class PinLoginRequest(BaseModel):
    pin: str

class EmailLoginRequest(BaseModel):
    email: str
    password: str

class SetPasswordRequest(BaseModel):
    new_password: str

class ResetRequestModel(BaseModel):
    email: str
    reason: str = ""

class TempPasswordRequest(BaseModel):
    email: str
    temp_password: str
    new_password: str

class UserRole(str, Enum):
    admin = "admin"
    full_access = "full_access"
    stock_only = "stock_only"
VALID_ROLES = {r.value for r in UserRole}

class CreateUserRequest(BaseModel):
    email: str
    display_name: str = ""
    pin: str = ""
    role: str = "stock_only"

class UpdateUserRequest(BaseModel):
    display_name: str = None
    pin: str = None
    role: str = None
    is_banned: bool = None

class TrackRequest(BaseModel):
    action: str
    details: dict = None


def call_baselinker(method: str, params: dict = None) -> dict:
    """Make BaseLinker API call"""
    if not BASELINKER_API_KEY:
        return {"error": "BASELINKER_API_KEY not configured"}

    data = {
        'token': BASELINKER_API_KEY,
        'method': method,
        'parameters': json.dumps(params or {})
    }
    try:
        response = requests.post(BASELINKER_API_URL, data=data, timeout=120)
        result = response.json()
        if result.get('status') == 'ERROR':
            return {"error": result.get('error_message', 'Unknown error')}
        return result
    except Exception as e:
        return {"error": str(e)}


def determine_category(name: str) -> str:
    """Determine product category from Polish product name"""
    name_lower = name.lower()

    # Coffee Tables (check BEFORE generic tables)
    if 'stolik kawowy' in name_lower or u'\u0142awa' in name_lower or 'lawa' in name_lower:
        return 'Coffee Tables'

    # Bar Stools (check BEFORE generic chairs)
    if 'hoker' in name_lower or 'barowy' in name_lower:
        return 'Bar Stools'

    # Ceramic Tables
    if 'ceramiczn' in name_lower or 'ceramic' in name_lower:
        if u'st\u00f3\u0142' in name_lower or 'stol' in name_lower or 'table' in name_lower:
            return 'Ceramic Tables'

    # Wooden Tables
    wood_keywords = ['drewn', 'industrialn', 'wooden', 'wood']
    if any(kw in name_lower for kw in wood_keywords):
        if u'st\u00f3\u0142' in name_lower or 'stol' in name_lower:
            if 'stolik' not in name_lower:
                return 'Wooden Tables'

    # Tables (generic, not caught above)
    if (u'st\u00f3\u0142' in name_lower or 'stol' in name_lower) and 'stolik' not in name_lower:
        return 'Tables'

    # Armchairs (check BEFORE generic chairs)
    if 'fotel' in name_lower or 'armchair' in name_lower:
        return 'Armchairs'

    # Chairs
    if u'krzes\u0142o' in name_lower or 'krzeslo' in name_lower or ('chair' in name_lower and 'armchair' not in name_lower):
        return 'Chairs'

    return 'Other'


# ========== COST LOADING ==========

def load_costs_and_pos_from_import_sheet():
    """Load product costs AND purchase order data from Google Sheets import tabs.

    Each tab = one invoice/purchase order. Tab name = invoice name.
    Returns (costs_by_sku, costs_by_base_sku, po_list, po_items_by_bl_id).
    PO matching is by BaseLinker Variant ID (column B) — exact numeric match, no SKU guessing.
    """
    costs_by_sku = {}
    costs_by_base_sku = {}
    po_list = []
    po_items_by_bl_id = {}  # BaseLinker Variant ID -> [PO references]

    if not GOOGLE_CREDENTIALS_JSON:
        logger.warning("GOOGLE_CREDENTIALS_JSON not configured, skipping cost loading")
        return costs_by_sku, costs_by_base_sku, po_list, po_items_by_bl_id

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
        credentials = Credentials.from_service_account_info(
            creds_dict,
            scopes=['https://www.googleapis.com/auth/spreadsheets.readonly']
        )
        client = gspread.authorize(credentials)
        sheet = client.open_by_key(IMPORT_SHEET_ID)

        for ws in sheet.worksheets():
            tab_name = ws.title
            all_data = ws.get_all_values()
            if len(all_data) < 9:
                continue

            # Find header row (scan first 15 rows)
            header_row_idx = None
            for i in range(min(15, len(all_data))):
                row_text = ' '.join([str(c).lower() for c in all_data[i][:25]])
                if 'model' in row_text and ('total cost' in row_text or 'unit price' in row_text):
                    header_row_idx = i
                    break

            if header_row_idx is None:
                continue

            headers = all_data[header_row_idx]

            # Dynamic column detection
            model_col = None
            color_col = None
            cost_col = None
            qty_col = None
            bl_variant_col = None

            for idx, header in enumerate(headers):
                h = str(header).lower().strip()
                if model_col is None and ('model' in h or h == 'sku'):
                    model_col = idx
                if color_col is None and ('color' in h or 'kolor' in h):
                    color_col = idx
                if cost_col is None and 'total cost' in h and 'pln' in h and 'all units' not in h:
                    cost_col = idx
                if qty_col is None and h in ('qty', 'quantity', 'ilosc', 'szt', 'ilość'):
                    qty_col = idx
                if bl_variant_col is None and ('baselinker' in h and 'variant' in h):
                    bl_variant_col = idx

            if model_col is None or cost_col is None:
                continue

            # Build PO entry for this tab
            po_entry = {
                'po_id': tab_name,
                'document_number': tab_name,
                'items_count': 0,
                'total_quantity': 0,
                'total_cost': 0.0,
                'product_skus': [],   # for SKU-based filtering
                'product_ids': [],    # BaseLinker variant IDs if available
            }

            count = 0
            for row in all_data[header_row_idx + 1:]:
                if len(row) <= max(model_col, cost_col):
                    continue

                model_raw = row[model_col].strip() if model_col < len(row) else ""
                color = row[color_col].strip().upper() if color_col and color_col < len(row) else ""
                cost_str = row[cost_col].strip() if cost_col < len(row) else ""
                qty_str = row[qty_col].strip() if qty_col is not None and qty_col < len(row) else "1"
                bl_variant = row[bl_variant_col].strip() if bl_variant_col is not None and bl_variant_col < len(row) else ""

                # Clean model: strip Chinese chars and other non-ASCII, normalize
                # e.g. "S-91定色451" → "S-91" (Chinese chars removed, trailing digits
                # belong to color column not model)
                model = re.sub(r'[^\x00-\x7F]+', '', model_raw).strip().upper()
                # If color column has a value and model ends with that color, strip it
                # e.g. model="S-91451" color="451" → model="S-91"
                if color and model.endswith(color):
                    model = model[:-len(color)].rstrip('-')

                if not model or not cost_str:
                    continue

                try:
                    cost = float(cost_str.replace('zł', '').replace('PLN', '').replace('$', '').replace(',', '').strip())
                except Exception:
                    continue

                try:
                    qty = int(float(qty_str)) if qty_str else 1
                except Exception:
                    qty = 1

                if cost <= 0:
                    continue

                # Build the full SKU key
                full_sku = f"{model}-{color}" if color and color not in ('N/A', '-', '') else model

                # Store costs
                if color and color not in ('N/A', '-', ''):
                    costs_by_sku[f"{model}-{color}"] = cost
                costs_by_base_sku[model] = cost

                # Build PO item reference
                po_item = {
                    'document_number': tab_name,
                    'quantity_ordered': qty,
                    'item_cost': cost,
                    'product_sku': full_sku,
                }

                # Add to PO items map keyed by BaseLinker Variant ID (exact match)
                if bl_variant and bl_variant != '0':
                    if bl_variant not in po_items_by_bl_id:
                        po_items_by_bl_id[bl_variant] = []
                    po_items_by_bl_id[bl_variant].append(po_item)

                # Track in PO entry
                po_entry['product_skus'].append(full_sku)
                po_entry['items_count'] += 1
                po_entry['total_quantity'] += qty
                po_entry['total_cost'] += cost * qty

                if bl_variant and bl_variant != '0':
                    po_entry['product_ids'].append(bl_variant)

                count += 1

            if count > 0:
                po_entry['total_cost'] = round(po_entry['total_cost'], 2)
                po_list.append(po_entry)
                logger.info(f"  Sheet '{tab_name}': {count} items, {po_entry['total_quantity']} units, {po_entry['total_cost']} PLN")

        logger.info(f"Total costs loaded: {len(costs_by_sku)} SKU, {len(costs_by_base_sku)} base SKU")
        logger.info(f"Total import POs: {len(po_list)} invoices, {len(po_items_by_bl_id)} BL variant ID mappings")

    except Exception as e:
        logger.error(f"Error loading costs/POs from sheets: {e}")
        import traceback
        traceback.print_exc()

    return costs_by_sku, costs_by_base_sku, po_list, po_items_by_bl_id


def find_cost_for_sku(sku, costs_by_sku, costs_by_base_sku):
    """Find cost for a SKU using multi-tier matching.
    Handles SKUs like S-91-451, DR2523-SZARY, DT-244-Czarny.
    Returns (cost, match_type).
    """
    if not sku:
        return 0, None

    sku_upper = sku.strip().upper()

    # 1. Exact match (full SKU including color suffix)
    if sku_upper in costs_by_sku:
        return costs_by_sku[sku_upper], 'exact'

    # 2. Progressive base SKU matching — try removing segments from the right
    # e.g. S-91-451 → try "S-91-451", "S-91", "S" against costs_by_base_sku
    parts = sku_upper.split('-')
    for i in range(len(parts) - 1, 0, -1):
        candidate = '-'.join(parts[:i])
        if candidate in costs_by_base_sku:
            return costs_by_base_sku[candidate], 'base'

    # Also try the full SKU as a base key (no-dash SKUs)
    if sku_upper in costs_by_base_sku:
        return costs_by_base_sku[sku_upper], 'base'

    # 3. Prefix match — find any cost key that starts with a base candidate
    for i in range(len(parts) - 1, 0, -1):
        candidate = '-'.join(parts[:i])
        for key, cost in costs_by_sku.items():
            if key.startswith(candidate + '-'):
                return cost, 'prefix'

    return 0, None


# ========== ORDER FETCHING ==========

def fetch_orders_by_status(status_id: int) -> list:
    """Fetch ALL orders with a given status ID."""
    all_orders = []
    last_order_id = 0

    while True:
        params = {
            'status_id': status_id,
            'get_unconfirmed_orders': False
        }
        if last_order_id > 0:
            params['id_from'] = last_order_id

        result = call_baselinker('getOrders', params)
        if "error" in result:
            break

        fetched_orders = result.get('orders', [])
        if not fetched_orders:
            break

        all_orders.extend(fetched_orders)
        last_order_id = fetched_orders[-1].get('order_id', 0)

        time.sleep(0.3)

        if len(fetched_orders) < 100:
            break

    return all_orders


def fetch_all_financial_orders() -> list:
    """Fetch ALL orders from statuses that count toward financials (Wysłane + Spakowane)."""
    all_orders = {}
    for status_id in FINANCIAL_STATUS_IDS:
        status_orders = fetch_orders_by_status(status_id)
        logger.info(f"Fetched {len(status_orders)} orders for status {status_id}")
        for o in status_orders:
            all_orders[o['order_id']] = o  # deduplicate by order_id
    logger.info(f"Total financial orders (Wysłane + Spakowane): {len(all_orders)}")
    return list(all_orders.values())


def fetch_recent_orders_all_statuses(days: int = 90) -> list:
    """Fetch orders from ALL statuses for date-filtered views.

    When filtering by date (Today, Yesterday, etc.), we want ALL orders placed
    on that date regardless of their current status — not just shipped ones.
    """
    from_ts = int((datetime.now(timezone.utc) - timedelta(days=days)).timestamp())
    all_orders = []
    last_order_id = 0

    while True:
        params = {
            'date_from': from_ts,
            'get_unconfirmed_orders': False,
        }
        # NO status_id param -> fetches all statuses
        if last_order_id > 0:
            params['id_from'] = last_order_id

        result = call_baselinker('getOrders', params)
        if "error" in result:
            logger.error(f"Error fetching all-status orders: {result.get('error')}")
            break

        fetched = result.get('orders', [])
        if not fetched:
            break

        all_orders.extend(fetched)
        last_order_id = fetched[-1].get('order_id', 0)
        time.sleep(0.3)

        if len(fetched) < 100:
            break

    logger.info(f"Fetched {len(all_orders)} orders from all statuses (last {days} days)")
    return all_orders


def fetch_order_sources() -> dict:
    """Fetch order sources mapping (source_id -> name) from BaseLinker API.

    API returns nested dict: {sources: {type: {account_id: name_or_info}}}
    e.g. {sources: {allegro: {24535: "Marbily", 26446: "Glamova"}, shop: {5017156: "Marbily"}}}
    """
    result = call_baselinker('getOrderSources')
    if "error" in result:
        logger.error(f"Error fetching order sources: {result.get('error')}, raw response: {result}")
        return {}

    source_names = {}

    # Iterate all top-level keys (skip 'status')
    for key, val in result.items():
        if key == 'status':
            continue
        # val is either the 'sources' wrapper dict or a direct type dict
        if not isinstance(val, dict):
            continue

        # Check if this is the 'sources' wrapper: {type: {id: name}}
        # or a direct type-level dict: {id: name}
        for source_type, accounts in val.items():
            if not isinstance(accounts, dict):
                continue
            # accounts = {account_id: name_string_or_dict}
            for acc_id, acc_name in accounts.items():
                if isinstance(acc_name, dict):
                    name = acc_name.get('name', str(acc_id))
                else:
                    name = str(acc_name)
                # Prefix with type for clarity (except personal/generic)
                type_label = source_type.capitalize()
                if type_label in ('Personal', 'Order_return', '0'):
                    source_names[str(acc_id)] = name
                else:
                    source_names[str(acc_id)] = f"{type_label} - {name}" if name != type_label else name

    logger.info(f"Parsed order sources: {source_names}")
    return source_names


def aggregate_sales(orders: list, source_names: dict = None) -> dict:
    """Aggregate sales by variant with revenue and channel tracking.

    Includes outlier detection: if a variant has 3+ price instances and one
    price is >10x the median, it's excluded from revenue (Allegro sometimes
    reports order totals as per-unit price).
    """
    sales_by_variant = defaultdict(lambda: {
        'product_name': '',
        'sku': '',
        'units_sold': 0,
        'bl_product_id': '',
        'total_revenue': 0.0,
        'sales_by_channel': defaultdict(int),
        '_price_instances': [],  # track individual prices for outlier detection
        '_revenue_by_currency': defaultdict(float),  # track revenue per currency for VAT calc
    })

    for order in orders:
        # Detect sales channel
        source = order.get('order_source', 'unknown')
        source_id = str(order.get('order_source_id', ''))

        # Use the source name from the mapping if available
        if source_names and source_id in source_names:
            channel = source_names[source_id]
        elif 'allegro' in source.lower():
            channel = 'Allegro'
        elif 'shopify' in source.lower() or source == 'shop':
            channel = 'Shopify'
        elif 'olx' in source.lower():
            channel = 'OLX'
        else:
            channel = source

        # Get order currency for price conversion
        order_currency = (order.get('currency', '') or 'PLN').upper()

        # Include delivery price in revenue (matches Marbily App logic)
        delivery_price_original = float(order.get('delivery_price', 0) or 0)
        delivery_price = convert_to_pln(delivery_price_original, order_currency)

        # First pass: collect products and their base revenue for proportional delivery split
        order_products = []
        order_products_total = 0.0
        for product in order.get('products', []):
            bl_product_id = str(product.get('variant_id', ''))
            sku = product.get('sku', '') or ''
            # If variant_id is 0/empty (e.g. Erli orders), use SKU as fallback identifier
            if (not bl_product_id or bl_product_id == '0') and not sku:
                continue  # skip only if BOTH variant_id and SKU are missing
            name = product.get('name', '') or 'Unknown'
            qty = int(product.get('quantity', 1))
            price_original = float(product.get('price_brutto') or 0)
            price = convert_to_pln(price_original, order_currency)
            product_revenue = price * qty
            order_products_total += product_revenue
            order_products.append({
                'bl_product_id': bl_product_id, 'sku': sku, 'name': name,
                'qty': qty, 'price': price, 'product_revenue': product_revenue,
            })

        # Second pass: distribute delivery fee proportionally and accumulate
        for p in order_products:
            variant_key = p['bl_product_id'] if p['bl_product_id'] and p['bl_product_id'] != '0' else p['sku']
            # Proportional share of delivery cost
            delivery_share = 0.0
            if delivery_price > 0:
                if order_products_total > 0:
                    delivery_share = delivery_price * (p['product_revenue'] / order_products_total)
                else:
                    delivery_share = delivery_price / len(order_products)

            sales_by_variant[variant_key]['product_name'] = p['name']
            sales_by_variant[variant_key]['sku'] = p['sku']
            sales_by_variant[variant_key]['units_sold'] += p['qty']
            sales_by_variant[variant_key]['bl_product_id'] = p['bl_product_id']
            sales_by_variant[variant_key]['total_revenue'] += p['product_revenue'] + delivery_share
            sales_by_variant[variant_key]['sales_by_channel'][channel] += p['qty']
            sales_by_variant[variant_key]['_price_instances'].append({'price': p['price'], 'qty': p['qty']})
            sales_by_variant[variant_key]['_revenue_by_currency'][order_currency] += p['product_revenue'] + delivery_share

    # Post-process: detect and exclude outlier prices per variant
    result = {}
    for key, val in sales_by_variant.items():
        instances = val.pop('_price_instances', [])
        prices = [inst['price'] for inst in instances if inst['price'] > 0]

        if len(prices) >= 3:
            sorted_prices = sorted(prices)
            median_price = sorted_prices[len(sorted_prices) // 2]
            if median_price > 0:
                # Recalculate revenue excluding outliers (price > 10x median)
                clean_revenue = sum(
                    inst['price'] * inst['qty'] for inst in instances
                    if inst['price'] <= median_price * 10
                )
                clean_units = sum(inst['qty'] for inst in instances if inst['price'] <= median_price * 10)
                if clean_revenue != val['total_revenue']:
                    excluded = val['total_revenue'] - clean_revenue
                    logger.warning(f"Outlier detected for {key}: excluded {excluded:.2f} PLN "
                          f"(median={median_price:.2f}, threshold={median_price*10:.2f})")
                    val['total_revenue'] = clean_revenue
                    val['units_sold'] = max(0, clean_units)

        # Floor units_sold and revenue at 0 — negative means returns exceeded sales
        val['units_sold'] = max(0, val['units_sold'])
        val['total_revenue'] = max(0.0, val['total_revenue'])

        val['sales_by_channel'] = dict(val['sales_by_channel'])
        val['_revenue_by_currency'] = dict(val.get('_revenue_by_currency', {}))
        result[key] = val

    return result




def fetch_full_inventory() -> list:
    """Fetch ALL products and variants from BaseLinker inventory for search.
    Returns a flat list of dicts with: bl_product_id, product_name, sku, current_stock, image_url, category, shopify_variant_id.
    """
    # Step 1: Get all parent products
    all_parents = {}
    page = 1
    while True:
        result = call_baselinker('getInventoryProductsList', {
            'inventory_id': BASELINKER_INVENTORY_ID,
            'page': page
        })
        if "error" in result:
            break
        products = result.get('products', {})
        if not products:
            break
        all_parents.update(products)
        page += 1
        time.sleep(0.3)

    logger.info(f"Inventory: {len(all_parents)} parent products found")

    # Step 2: Get detailed data for all parents (includes variants via getInventoryProductsData)
    parent_ids = list(int(pid) for pid in all_parents.keys())
    full_list = []
    # Collect Shopify product IDs to resolve variant IDs later
    shopify_product_ids = []  # list of Shopify product_id strings

    for i in range(0, len(parent_ids), 100):
        batch = parent_ids[i:i+100]
        result = call_baselinker('getInventoryProductsData', {
            'inventory_id': BASELINKER_INVENTORY_ID,
            'products': batch
        })
        if "error" in result:
            continue

        for pid_str, pdata in result.get('products', {}).items():
            name = pdata.get('text_fields', {}).get('name', '') or all_parents.get(pid_str, {}).get('name', '')
            sku = pdata.get('text_fields', {}).get('sku', '') or all_parents.get(pid_str, {}).get('sku', '')
            stock_data = pdata.get('stock', {})
            total_stock = sum(int(s) for s in stock_data.values() if s is not None)
            images = pdata.get('images', {})
            image_url = list(images.values())[0] if images else ''

            # Get Shopify product ID from links
            product_links = pdata.get('links', {})
            shopify_link = product_links.get('shop_5017156', {}) or {}
            shopify_pid = str(shopify_link.get('product_id', '')) if isinstance(shopify_link, dict) else ''
            if shopify_pid and shopify_pid != '0':
                shopify_product_ids.append(shopify_pid)

            # Check if this product has variants
            variants = pdata.get('variants', {})
            if variants:
                for vid, vdata in variants.items():
                    v_name_raw = vdata.get('name', '')
                    v_sku = vdata.get('sku', '') or sku
                    v_stock_data = vdata.get('stock', {})
                    v_stock = sum(int(s) for s in v_stock_data.values() if s is not None) if v_stock_data is not None else total_stock
                    v_images = vdata.get('images', {})
                    v_image = list(v_images.values())[0] if v_images else image_url
                    v_display = f"{name} - {v_name_raw}" if v_name_raw else name

                    full_list.append({
                        'bl_product_id': vid,
                        'product_name': v_display,
                        'sku': v_sku,
                        'current_stock': v_stock,
                        'image_url': v_image,
                        'category': determine_category(name),
                        'shopify_variant_id': '',  # filled in Step 3
                        'units_sold': 0,
                        'total_revenue': 0,
                        'sales_by_channel': {},
                        'purchase_orders': [],
                    })
            else:
                full_list.append({
                    'bl_product_id': pid_str,
                    'product_name': name,
                    'sku': sku,
                    'current_stock': total_stock,
                    'image_url': image_url,
                    'category': determine_category(name),
                    'shopify_variant_id': '',  # filled in Step 3
                    'units_sold': 0,
                    'total_revenue': 0,
                    'sales_by_channel': {},
                    'purchase_orders': [],
                })

        time.sleep(0.5)

    # Step 3: Resolve Shopify variant IDs by fetching external storage data and matching by SKU
    # Build SKU -> full_list index for fast lookup
    sku_to_items = defaultdict(list)
    for item in full_list:
        if item.get('sku'):
            sku_to_items[item['sku']].append(item)

    unique_shopify_pids = list(set(shopify_product_ids))
    logger.info(f"Resolving Shopify variant IDs for {len(unique_shopify_pids)} linked products...")
    matched = 0
    for i in range(0, len(unique_shopify_pids), 50):
        batch = [int(pid) for pid in unique_shopify_pids[i:i+50]]
        result = call_baselinker('getExternalStorageProductsData', {
            'storage_id': 'shop_5017156',
            'products': batch
        })
        if "error" in result:
            continue
        for spid_str, sdata in result.get('products', {}).items():
            for sv in sdata.get('variants', []):
                sv_sku = sv.get('sku', '')
                sv_vid = str(sv.get('variant_id', ''))
                if sv_sku and sv_vid and sv_sku in sku_to_items:
                    for item in sku_to_items[sv_sku]:
                        if not item.get('shopify_variant_id'):  # only assign if not already set
                            item['shopify_variant_id'] = sv_vid
                            matched += 1
                            break  # one Shopify variant per SKU match
        time.sleep(0.5)

    logger.info(f"Full inventory: {len(full_list)} products/variants indexed, {matched} Shopify variant IDs matched by SKU")

    # Deduplicate by SKU — keep entry with highest stock per SKU
    seen_skus = {}
    for item in full_list:
        sku = item.get('sku', '')
        if not sku:
            continue
        existing = seen_skus.get(sku)
        if not existing or item.get('current_stock', 0) > existing.get('current_stock', 0):
            seen_skus[sku] = item
    deduped = [item for item in full_list if not item.get('sku') or seen_skus.get(item.get('sku')) is item]
    logger.info(f"Deduplicated inventory: {len(full_list)} -> {len(deduped)} (removed {len(full_list) - len(deduped)} duplicate SKUs)")
    full_list = deduped

    return full_list


def filter_products(products: list, category: str = "", po: str = "", search: str = "") -> list:
    """Apply filters to product list (shared between API and Excel download)"""
    filtered = products

    if category:
        filtered = [p for p in filtered if p.get('category') == category]

    if po:
        po_data = next((p for p in cache.get("purchase_orders", []) if str(p.get('po_id')) == po), None)
        if po_data and po_data.get('product_ids'):
            # Filter by BaseLinker Variant IDs — exact match, no guessing
            bl_id_set = set(str(bid) for bid in po_data['product_ids'])
            filtered = [p for p in filtered if str(p.get('bl_product_id', '')) in bl_id_set]

    if search:
        variant_match = re.search(r'[?&]variant=(\d+)', search)
        if variant_match:
            vid = variant_match.group(1)
            filtered = [p for p in filtered if p.get('shopify_variant_id') == vid]
        else:
            q = search.lower()
            filtered = [p for p in filtered if
                q in (p.get('product_name', '') or '').lower() or
                q in (p.get('sku', '') or '').lower() or
                q in (p.get('bl_product_id', '') or '') or
                q in (p.get('category', '') or '').lower()
            ]

    return filtered


def build_response(orders: list, inventory: dict, po_items_map: dict, source_names: dict = None, costs: tuple = None) -> dict:
    """Build product list from orders + inventory + PO data. Reusable for date-filtered views."""
    sales_data = aggregate_sales(orders, source_names)

    products = []
    total_revenue = 0.0
    po_match_count = 0
    global_channel_units = defaultdict(int)
    global_channel_revenue = defaultdict(float)

    # Build SKU fallback index from full_inventory cache
    # Keep the entry with highest stock per SKU (catalog duplicates have stale/negative stock)
    _sku_inv_fallback = {}
    for item in cache.get("full_inventory", []):
        sku = item.get('sku', '')
        if sku:
            entry = {
                'stock': item.get('current_stock', 0),
                'image_url': item.get('image_url', ''),
                'shopify_variant_id': item.get('shopify_variant_id', ''),
            }
            existing = _sku_inv_fallback.get(sku)
            if not existing or entry['stock'] > existing['stock']:
                _sku_inv_fallback[sku] = entry

    for variant_key, data in sales_data.items():
        bl_pid = data['bl_product_id']
        inv_info = inventory.get(bl_pid, {})
        # Fallback: if stock is zero/negative, get ONLY the stock from the real inventory
        # product (matched by SKU). Keep the original image from the catalog product —
        # catalog products have the correct product-specific images.
        if data['sku'] and inv_info.get('stock', 0) <= 0:
            sku_fallback = _sku_inv_fallback.get(data['sku'], {})
            if sku_fallback.get('stock', 0) > 0:
                inv_info = {
                    'stock': sku_fallback['stock'],
                    'image_url': inv_info.get('image_url') or sku_fallback.get('image_url', ''),
                    'shopify_variant_id': inv_info.get('shopify_variant_id') or sku_fallback.get('shopify_variant_id', ''),
                }
        elif not inv_info and data['sku']:
            inv_info = _sku_inv_fallback.get(data['sku'], {})
        revenue = round(data['total_revenue'], 2)
        total_revenue += revenue
        # Match POs by BaseLinker Variant ID — exact numeric match from import sheet column B
        product_pos = po_items_map.get(bl_pid, []) if po_items_map else []
        if product_pos:
            po_match_count += 1

        # Floor units_sold at 0 (negative quantities from returns get netted)
        units_sold = max(0, data['units_sold'])

        # Stock: keep raw value but flag negative
        raw_stock = max(0, inv_info.get('stock', 0))

        products.append({
            'image_url': inv_info.get('image_url', ''),
            'product_name': data['product_name'],
            'sku': data['sku'],
            'units_sold': units_sold,
            'current_stock': raw_stock,
            'total_revenue': revenue,
            'sales_by_channel': data['sales_by_channel'],
            'category': determine_category(data['product_name']),
            'shopify_variant_id': inv_info.get('shopify_variant_id', ''),
            'bl_product_id': bl_pid,
            'purchase_orders': product_pos,
            '_revenue_by_currency': data.get('_revenue_by_currency', {}),
        })

        # Accumulate global channel breakdown
        for channel, qty in data['sales_by_channel'].items():
            global_channel_units[channel] += max(0, qty)
            # Estimate revenue per channel proportionally
        # Revenue by channel: distribute proportionally by qty
        total_qty = sum(data['sales_by_channel'].values())
        if total_qty > 0:
            for channel, qty in data['sales_by_channel'].items():
                if qty > 0:
                    global_channel_revenue[channel] += revenue * (qty / total_qty)

    products.sort(key=lambda x: x['units_sold'], reverse=True)

    # ========== COST ENRICHMENT ==========
    costs_by_sku, costs_by_base_sku = costs if costs else cache.get("costs", ({}, {}))
    total_cost = 0.0

    for product in products:
        cost, match_type = find_cost_for_sku(product['sku'], costs_by_sku, costs_by_base_sku)
        if cost == 0 and product['units_sold'] > 0 and product['total_revenue'] > 0:
            avg_price = product['total_revenue'] / product['units_sold']
            cost = avg_price / FALLBACK_MARKUP_GROSS
            match_type = 'fallback'
        product['unit_cost'] = round(cost, 2)
        product['total_cost'] = round(cost * product['units_sold'], 2)
        product['cost_match'] = match_type
        total_cost += product['total_cost']

    # Financial calculations — per-currency VAT rates for accurate net revenue
    # Aggregate revenue by currency across all products
    total_revenue_by_currency = defaultdict(float)
    for product in products:
        for curr, rev in product.get('_revenue_by_currency', {}).items():
            total_revenue_by_currency[curr] += rev
    # Compute blended net revenue using per-currency VAT rates
    net_revenue = 0.0
    for curr, rev in total_revenue_by_currency.items():
        vat_rate = VAT_RATES_BY_CURRENCY.get(curr, DEFAULT_VAT_RATE)
        net_revenue += rev / (1 + vat_rate) if vat_rate > 0 else rev
    # Fallback if no currency data available
    if net_revenue == 0 and total_revenue > 0:
        net_revenue = total_revenue / 1.23
    net_revenue = round(net_revenue, 2)
    profit = round(net_revenue - total_cost, 2)
    profit_margin = round((profit / net_revenue) * 100, 1) if net_revenue > 0 else 0

    # Build sorted channel breakdown (highest units first)
    channel_breakdown = []
    for ch in sorted(global_channel_units.keys(), key=lambda c: global_channel_units[c], reverse=True):
        channel_breakdown.append({
            "channel": ch,
            "units": global_channel_units[ch],
            "revenue": round(global_channel_revenue.get(ch, 0), 2)
        })

    return {
        "total_variants": len(products),
        "total_units_sold": sum(p['units_sold'] for p in products),
        "total_orders": len(orders),
        "total_revenue": round(total_revenue, 2),
        "net_revenue": net_revenue,
        "total_cost": round(total_cost, 2),
        "profit": profit,
        "profit_margin": profit_margin,
        "channel_breakdown": channel_breakdown,
        "exchange_rates_as_of": EXCHANGE_RATES_AS_OF,
        "products": products
    }


_refresh_lock = threading.Lock()

def refresh_data():
    """Fetch fresh data from BaseLinker"""
    if not _refresh_lock.acquire(blocking=False):
        logger.info("Refresh already in progress, skipping")
        return cache["data"]

    cache["is_refreshing"] = True

    try:
        # Fetch order sources for proper channel names
        order_sources = fetch_order_sources()
        cache["order_sources"] = order_sources

        # Load costs AND purchase orders from Google Sheets import tabs
        # Each tab = one invoice. Tab name = invoice name.
        # Wrapped separately so a Google Sheets error doesn't block order fetching
        po_list = []
        po_items_by_bl_id = {}
        try:
            logger.info("Loading costs and purchase orders from import sheets...")
            costs_by_sku, costs_by_base_sku, po_list, po_items_by_bl_id = load_costs_and_pos_from_import_sheet()
            cache["costs"] = (costs_by_sku, costs_by_base_sku)
            cache["po_items_by_bl_id"] = po_items_by_bl_id
        except Exception as e:
            logger.warning(f"Cost/PO loading failed (orders will still load): {e}")
            import traceback
            traceback.print_exc()

        # Fetch orders and aggregate sales
        orders = fetch_all_financial_orders()

        # Also fetch recent orders from ALL statuses for date-filtered views
        # (Today/Yesterday/etc. should show all orders, not just shipped ones)
        all_recent = fetch_recent_orders_all_statuses(days=90)

        # Save ALL fetched orders to the database (survives BaseLinker's 3-month archive)
        all_fetched = {o['order_id']: o for o in orders}
        for o in all_recent:
            all_fetched[o['order_id']] = o  # all_recent may have fresher status
        save_orders_to_db(list(all_fetched.values()))

        # For "All Time" view: merge fresh API orders with historical DB orders
        # DB has orders that BaseLinker may have archived (older than 3 months)
        db_orders = load_orders_from_db()
        merged_financial = {}
        # Start with DB orders (includes historical) — only financial statuses
        for o in db_orders:
            o_status = o.get('order_status_id') if o.get('order_status_id') is not None else o.get('status_id')
            if o_status in FINANCIAL_STATUS_IDS:
                merged_financial[o['order_id']] = o
        # Override with fresh API data (more up-to-date)
        for o in orders:
            merged_financial[o['order_id']] = o
        orders_merged = list(merged_financial.values())

        db_order_count = get_db_order_count()
        api_order_count = len(orders)
        logger.info(f"Orders: {api_order_count} from API, {db_order_count} in DB, {len(orders_merged)} merged for All Time view")

        # Cache raw orders — merged includes DB historical
        cache["raw_orders"] = orders_merged
        cache["all_recent_orders"] = all_recent

        sales_data = aggregate_sales(orders_merged, order_sources)
        product_ids = set(d['bl_product_id'] for d in sales_data.values() if d['bl_product_id'])

        # Also include product IDs from recent all-status orders
        # so that Today/Yesterday products have images even if not yet shipped
        for order in all_recent:
            for product in order.get('products', []):
                vid = str(product.get('variant_id', ''))
                if vid and vid != '0':
                    product_ids.add(vid)

        # Fetch full inventory for search (all products/variants with names, SKUs, stock)
        try:
            full_inv = fetch_full_inventory()
            cache["full_inventory"] = full_inv
        except Exception as e:
            logger.warning(f"Full inventory fetch failed (search may be limited): {e}")

        # Build inventory lookup from full_inventory.
        # CRITICAL: Orders use CATALOG variant_ids (e.g. 478011847) which are DIFFERENT
        # from INVENTORY variant_ids (e.g. 503940001). They share the same SKU.
        # So we build TWO indexes: by inventory variant ID AND by SKU.
        full_inv_data = cache.get("full_inventory", [])
        inventory = {}
        sku_to_inv = {}  # SKU -> inventory data (for matching orders by SKU)

        for item in full_inv_data:
            vid = str(item.get('bl_product_id', ''))
            inv_entry = {
                'stock': item.get('current_stock', 0),
                'image_url': item.get('image_url', ''),
                'shopify_variant_id': item.get('shopify_variant_id', ''),
            }
            if vid:
                # If this vid already exists, keep the entry with a real image
                # (variants inherit parent images which may be wrong for the specific variant)
                existing_inv = inventory.get(vid)
                if not existing_inv or (inv_entry['image_url'] and inv_entry['stock'] >= existing_inv.get('stock', 0)):
                    inventory[vid] = inv_entry
                elif inv_entry['stock'] > existing_inv.get('stock', 0):
                    # Better stock but maybe worse image — only update stock
                    existing_inv['stock'] = inv_entry['stock']
            sku = item.get('sku', '')
            if sku:
                # Keep the entry with highest stock — catalog duplicates often show
                # negative/stale stock while the real inventory product has correct stock
                existing = sku_to_inv.get(sku)
                if not existing or inv_entry['stock'] > existing['stock']:
                    sku_to_inv[sku] = inv_entry

        logger.info(f"Inventory indexed: {len(inventory)} by variant ID, {len(sku_to_inv)} by SKU")

        # Also index by catalog variant_id from orders (product_ids set)
        # These are the IDs that build_response will look up
        # Match them to inventory via SKU from the orders
        sales_data = aggregate_sales(orders_merged, order_sources)
        sku_matched = 0
        for variant_key, data in sales_data.items():
            bl_pid = data['bl_product_id']
            sku = data['sku']
            # If catalog variant_id not in inventory OR has bad stock, use SKU match
            if bl_pid and sku and sku in sku_to_inv:
                existing = inventory.get(bl_pid)
                sku_entry = sku_to_inv[sku]
                if not existing:
                    inventory[bl_pid] = sku_entry
                    sku_matched += 1
                elif existing.get('stock', 0) <= 0 and sku_entry.get('stock', 0) > 0:
                    # Only override stock, keep original image from catalog product
                    existing['stock'] = sku_entry['stock']
                    if not existing.get('shopify_variant_id'):
                        existing['shopify_variant_id'] = sku_entry.get('shopify_variant_id', '')
                    sku_matched += 1
        if sku_matched:
            logger.info(f"Inventory SKU-matched: {sku_matched} catalog variant IDs linked to inventory via SKU")

        # Fetch images for order product_ids directly from BaseLinker API.
        # The catalog variant_ids from orders return their own correct images when
        # queried as standalone products. full_inventory may return parent images instead.
        unique_pids = list(set(int(pid) for pid in product_ids if pid))
        img_fixed = 0
        for i in range(0, len(unique_pids), 100):
            batch = unique_pids[i:i+100]
            try:
                result = call_baselinker('getInventoryProductsData', {
                    'inventory_id': BASELINKER_INVENTORY_ID,
                    'products': batch
                })
                for pid_str, product_info in result.get('products', {}).items():
                    images = product_info.get('images', {})
                    if images and pid_str in inventory:
                        real_img = list(images.values())[0]
                        if inventory[pid_str].get('image_url') != real_img:
                            inventory[pid_str]['image_url'] = real_img
                            img_fixed += 1
                time.sleep(0.5)
            except Exception:
                pass
        if img_fixed:
            logger.info(f"Images corrected: {img_fixed} product images updated from direct API lookup")

        cache["inventory"] = inventory

        # PO items map for build_response (keyed by BaseLinker Variant ID)
        po_items_map = po_items_by_bl_id

        # Build full response using merged orders (API + DB historical)
        response_data = build_response(orders_merged, inventory, po_items_map, order_sources)

        now = datetime.now(timezone.utc)
        cache["data"] = response_data
        cache["last_updated"] = now.isoformat()
        cache["next_refresh"] = (now + timedelta(seconds=REFRESH_INTERVAL)).isoformat()
        cache["purchase_orders"] = po_list

        # Log to database
        log_activity("refresh", cache["data"])
        save_sales_snapshot(response_data["products"])

    except Exception as e:
        logger.error(f"Error refreshing data: {e}")
        import traceback
        traceback.print_exc()
        log_activity("error", {"message": str(e)})
    finally:
        cache["is_refreshing"] = False
        _refresh_lock.release()

    return cache["data"]


async def background_refresh_task():
    """Background task to refresh data periodically"""
    while True:
        await asyncio.sleep(REFRESH_INTERVAL)
        if not cache["is_refreshing"]:
            await asyncio.to_thread(refresh_data)


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup and shutdown events"""
    if not BASELINKER_API_KEY:
        logger.error("CRITICAL: BASELINKER_API_KEY not set!")
    if not DATABASE_URL:
        logger.error("CRITICAL: DATABASE_URL not set!")
    init_db_pool()
    init_database()
    log_activity("startup")
    # Run initial refresh in background thread — don't block startup
    # so Railway health checks pass while data loads
    loop = asyncio.get_event_loop()
    loop.run_in_executor(None, refresh_data)
    task = asyncio.create_task(background_refresh_task())
    yield
    task.cancel()


app = FastAPI(
    title="BaseLinker Sales Dashboard",
    description="Real-time sales tracking from BaseLinker",
    lifespan=lifespan
)

app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
async def root():
    return FileResponse("static/index.html")


@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "last_updated": cache["last_updated"],
        "data_loaded": cache["data"] is not None,
        "is_refreshing": cache["is_refreshing"],
        "database_connected": db_pool is not None,
        "orders_in_db": get_db_order_count()
    }


# ========== AUTH ROUTES ==========

@app.post("/api/auth/pin")
async def pin_login(req: PinLoginRequest, request: Request):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")

    try:
        ip = get_client_ip(request)

        # Rate limit: 5 attempts per IP in 5 minutes — fail-closed on DB error
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT COUNT(*) FROM pin_attempts
                WHERE ip_address = %s AND attempted_at > NOW() - INTERVAL '5 minutes' AND NOT success
            """, (ip,))
            attempts = cur.fetchone()[0]
            cur.close()

            if attempts >= 5:
                return {"success": False, "fallback": "email_password", "message": "Too many PIN attempts. Use email and password."}
        except Exception:
            try: conn.rollback()
            except Exception: pass
            raise HTTPException(status_code=503, detail="Rate limit check unavailable")

        # Find users with PINs
        try:
            cur = conn.cursor()
            cur.execute("SELECT id, email, display_name, role, pin_hash, is_banned, must_change_password FROM users WHERE pin_hash IS NOT NULL")
            users = cur.fetchall()
            cur.close()
        except Exception as e:
            try: conn.rollback()
            except Exception: pass
            raise HTTPException(status_code=500, detail=str(e))

        matched_user = None
        for u in users:
            if u[4] and verify_pin(req.pin, u[4]):
                matched_user = u
                break

        # Log attempt
        try:
            cur = conn.cursor()
            cur.execute("INSERT INTO pin_attempts (ip_address, success) VALUES (%s, %s)", (ip, matched_user is not None))
            conn.commit()
            cur.close()
        except Exception:
            try: conn.rollback()
            except Exception: pass

        if not matched_user:
            return {"success": False, "message": "Invalid PIN"}

        if matched_user[5]:  # is_banned
            return {"success": False, "message": "Account is banned"}

        # Clear failed attempts on successful login
        try:
            cur = conn.cursor()
            cur.execute("DELETE FROM pin_attempts WHERE ip_address = %s", (ip,))
            conn.commit()
            cur.close()
        except Exception:
            try: conn.rollback()
            except Exception: pass

        token = create_session(matched_user[0], request)
        log_user_activity(matched_user[0], "", "pin_login", {"ip": ip})

        return {
            "success": True,
            "token": token,
            "user": {
                "id": matched_user[0],
                "email": matched_user[1],
                "display_name": matched_user[2],
                "role": matched_user[3],
                "must_change_password": matched_user[6]
            }
        }
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.post("/api/auth/login")
async def email_login(req: EmailLoginRequest, request: Request):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")

    try:
        try:
            cur = conn.cursor()
            cur.execute("SELECT id, email, display_name, role, password_hash, is_banned, must_change_password FROM users WHERE email = %s", (req.email.lower().strip(),))
            user = cur.fetchone()
            cur.close()
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

        if not user or not user[4]:
            return {"success": False, "message": "Invalid email or password"}

        if not verify_password(req.password, user[4]):
            return {"success": False, "message": "Invalid email or password"}

        if user[5]:  # is_banned
            return {"success": False, "message": "Account is banned"}

        token = create_session(user[0], request)
        ip = get_client_ip(request)
        log_user_activity(user[0], "", "email_login", {"ip": ip})

        return {
            "success": True,
            "token": token,
            "user": {
                "id": user[0],
                "email": user[1],
                "display_name": user[2],
                "role": user[3],
                "must_change_password": user[6]
            }
        }
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.post("/api/auth/set-password")
async def set_password(req: SetPasswordRequest, user: dict = Depends(get_current_user)):
    if len(req.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")

    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")

    try:
        cur = conn.cursor()
        pw_hash = hash_password(req.new_password)
        cur.execute("UPDATE users SET password_hash = %s, must_change_password = FALSE, updated_at = NOW() WHERE id = %s", (pw_hash, user["id"]))
        conn.commit()
        cur.close()
        log_user_activity(user["id"], user.get("jti", ""), "set_password", {})
        return {"success": True, "message": "Password set successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.post("/api/auth/reset-request")
async def request_reset(req: ResetRequestModel):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")

    try:
        cur = conn.cursor()
        cur.execute("SELECT id FROM users WHERE email = %s", (req.email.lower().strip(),))
        user = cur.fetchone()
        if not user:
            return {"success": True, "message": "If an account exists, your request has been sent to the admin."}

        cur.execute("""
            INSERT INTO password_reset_requests (user_id, reason) VALUES (%s, %s)
        """, (user[0], req.reason))
        conn.commit()
        cur.close()
        return {"success": True, "message": "Your request has been sent to the admin."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.post("/api/auth/use-temp-password")
async def use_temp_password(req: TempPasswordRequest, request: Request):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")

    try:
        cur = conn.cursor()
        cur.execute("SELECT id, display_name, role FROM users WHERE email = %s", (req.email.lower().strip(),))
        user = cur.fetchone()
        if not user:
            return {"success": False, "message": "Invalid credentials"}

        # Check temp password in reset requests
        cur.execute("""
            SELECT id, temp_password FROM password_reset_requests
            WHERE user_id = %s AND status = 'fulfilled' AND temp_password IS NOT NULL
            ORDER BY fulfilled_at DESC LIMIT 1
        """, (user[0],))
        reset_req = cur.fetchone()

        if not reset_req or not verify_password(req.temp_password, reset_req[1]):
            return {"success": False, "message": "Invalid temp password"}

        # Set new password and mark reset as used
        pw_hash = hash_password(req.new_password)
        cur.execute("UPDATE users SET password_hash = %s, must_change_password = FALSE, updated_at = NOW() WHERE id = %s", (pw_hash, user[0]))
        cur.execute("UPDATE password_reset_requests SET status = 'used', temp_password = NULL WHERE id = %s", (reset_req[0],))
        conn.commit()

        token = create_session(user[0], request)
        cur.close()

        return {
            "success": True,
            "token": token,
            "user": {"id": user[0], "email": req.email, "display_name": user[1], "role": user[2], "must_change_password": False}
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.post("/api/auth/logout")
async def logout(user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE user_sessions SET is_active = FALSE WHERE token_jti = %s", (user.get("jti"),))
            conn.commit()
            cur.close()
        except Exception:
            pass
        finally:
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass
            put_db_connection(conn)
    log_user_activity(user["id"], user.get("jti", ""), "logout", {})
    return {"success": True}

@app.get("/api/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return {"user": user}

@app.post("/api/auth/heartbeat")
async def heartbeat(user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE user_sessions SET last_heartbeat = NOW() WHERE token_jti = %s", (user.get("jti"),))
            conn.commit()
            cur.close()
        except Exception:
            pass
        finally:
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass
            put_db_connection(conn)
    return {"ok": True}


# ========== ADMIN ROUTES ==========

@app.get("/api/admin/users")
async def admin_list_users(user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        return {"users": []}
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, email, display_name, role, is_banned, must_change_password, created_at FROM users ORDER BY id")
        rows = cur.fetchall()
        cur.close()
        return {"users": [{"id": r[0], "email": r[1], "display_name": r[2], "role": r[3], "is_banned": r[4], "must_change_password": r[5], "created_at": r[6].isoformat() if r[6] else None} for r in rows]}
    except Exception as e:
        return {"error": str(e), "users": []}
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.post("/api/admin/users")
async def admin_create_user(req: CreateUserRequest, user: dict = Depends(require_admin)):
    if req.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {', '.join(VALID_ROLES)}")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")
    try:
        cur = conn.cursor()
        pin_h = bcrypt.hashpw(req.pin.encode(), bcrypt.gensalt()).decode() if req.pin else None
        cur.execute("""
            INSERT INTO users (email, display_name, pin_hash, role, must_change_password)
            VALUES (%s, %s, %s, %s, TRUE) RETURNING id
        """, (req.email.lower().strip(), req.display_name, pin_h, req.role))
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        log_user_activity(user["id"], user.get("jti", ""), "admin_create_user", {"new_user_id": new_id, "email": req.email})
        return {"success": True, "user_id": new_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.put("/api/admin/users/{user_id}")
async def admin_update_user(user_id: int, req: UpdateUserRequest, user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")
    try:
        cur = conn.cursor()
        updates = []
        params = []
        if req.display_name is not None:
            updates.append("display_name = %s")
            params.append(req.display_name)
        if req.pin is not None and req.pin:
            updates.append("pin_hash = %s")
            params.append(bcrypt.hashpw(req.pin.encode(), bcrypt.gensalt()).decode())
        if req.role is not None:
            if req.role not in VALID_ROLES:
                raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {', '.join(VALID_ROLES)}")
            updates.append("role = %s")
            params.append(req.role)
        if req.is_banned is not None:
            updates.append("is_banned = %s")
            params.append(req.is_banned)
        if updates:
            updates.append("updated_at = NOW()")
            params.append(user_id)
            cur.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = %s", params)
            conn.commit()
        cur.close()
        log_user_activity(user["id"], user.get("jti", ""), "admin_update_user", {"target_user_id": user_id})
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.delete("/api/admin/users/{user_id}")
async def admin_delete_user(user_id: int, user: dict = Depends(require_admin)):
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
        conn.commit()
        cur.close()
        log_user_activity(user["id"], user.get("jti", ""), "admin_delete_user", {"deleted_user_id": user_id})
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.get("/api/admin/sessions")
async def admin_sessions(user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        return {"sessions": []}
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT s.id, u.email, u.display_name, s.created_at, s.last_heartbeat, s.ip_address, s.user_agent,
                   (s.last_heartbeat > NOW() - INTERVAL '5 minutes') as is_online
            FROM user_sessions s
            JOIN users u ON u.id = s.user_id
            WHERE s.is_active = TRUE AND s.expires_at > NOW()
            ORDER BY s.last_heartbeat DESC
        """)
        rows = cur.fetchall()
        cur.close()
        return {"sessions": [{"id": r[0], "email": r[1], "display_name": r[2], "created_at": r[3].isoformat() if r[3] else None, "last_heartbeat": r[4].isoformat() if r[4] else None, "ip_address": r[5], "user_agent": r[6], "is_online": r[7]} for r in rows]}
    except Exception as e:
        return {"error": str(e), "sessions": []}
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.get("/api/admin/screen-time")
async def admin_screen_time(user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        return {"screen_time": []}
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT u.email, u.display_name,
                   DATE(s.created_at AT TIME ZONE 'Europe/Warsaw') as session_date,
                   LEAST(SUM(LEAST(EXTRACT(EPOCH FROM (s.last_heartbeat - s.created_at)), 86400)) / 60, 1440) as minutes
            FROM user_sessions s
            JOIN users u ON u.id = s.user_id
            WHERE s.last_heartbeat > s.created_at
            GROUP BY u.email, u.display_name, DATE(s.created_at AT TIME ZONE 'Europe/Warsaw')
            ORDER BY session_date DESC, minutes DESC
            LIMIT 100
        """)
        rows = cur.fetchall()
        cur.close()
        return {"screen_time": [{"email": r[0], "display_name": r[1], "date": r[2].isoformat() if r[2] else None, "minutes": round(float(r[3] or 0), 1)} for r in rows]}
    except Exception as e:
        return {"error": str(e), "screen_time": []}
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.get("/api/admin/search-history")
async def admin_search_history(user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        return {"searches": []}
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT u.email, ua.timestamp, ua.details->>'query' as query, ua.details->>'filters' as filters
            FROM user_activity ua
            JOIN users u ON u.id = ua.user_id
            WHERE ua.action IN ('search', 'filter', 'view_sales')
            ORDER BY ua.timestamp DESC
            LIMIT 200
        """)
        rows = cur.fetchall()
        cur.close()
        return {"searches": [{"email": r[0], "timestamp": r[1].isoformat() if r[1] else None, "query": r[2], "filters": r[3]} for r in rows]}
    except Exception as e:
        return {"error": str(e), "searches": []}
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.get("/api/admin/activity")
async def admin_activity_log(page: int = Query(1, ge=1), limit: int = Query(50, ge=1, le=200), user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        return {"activities": [], "total": 0}
    try:
        cur = conn.cursor()
        offset = (page - 1) * limit
        cur.execute("SELECT COUNT(*) FROM user_activity")
        total = cur.fetchone()[0]
        cur.execute("""
            SELECT ua.id, u.email, u.display_name, ua.timestamp, ua.action, ua.details
            FROM user_activity ua
            JOIN users u ON u.id = ua.user_id
            ORDER BY ua.timestamp DESC
            LIMIT %s OFFSET %s
        """, (limit, offset))
        rows = cur.fetchall()
        cur.close()
        return {
            "activities": [{"id": r[0], "email": r[1], "display_name": r[2], "timestamp": r[3].isoformat() if r[3] else None, "action": r[4], "details": r[5]} for r in rows],
            "total": total,
            "page": page,
            "pages": (total + limit - 1) // limit
        }
    except Exception as e:
        return {"error": str(e), "activities": [], "total": 0}
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.get("/api/admin/reset-requests")
async def admin_reset_requests(user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        return {"requests": []}
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT pr.id, u.email, u.display_name, pr.requested_at, pr.reason, pr.status
            FROM password_reset_requests pr
            JOIN users u ON u.id = pr.user_id
            ORDER BY pr.requested_at DESC
            LIMIT 50
        """)
        rows = cur.fetchall()
        cur.close()
        return {"requests": [{"id": r[0], "email": r[1], "display_name": r[2], "requested_at": r[3].isoformat() if r[3] else None, "reason": r[4], "status": r[5]} for r in rows]}
    except Exception as e:
        return {"error": str(e), "requests": []}
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.post("/api/admin/reset-requests/{request_id}/fulfill")
async def admin_fulfill_reset(request_id: int, user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")
    try:
        temp_pw = secrets.token_urlsafe(8)
        temp_hash = hash_password(temp_pw)
        cur = conn.cursor()
        cur.execute("""
            UPDATE password_reset_requests
            SET status = 'fulfilled', fulfilled_at = NOW(), fulfilled_by = %s, temp_password = %s
            WHERE id = %s AND status = 'pending'
        """, (user["id"], temp_hash, request_id))
        conn.commit()
        cur.close()
        log_user_activity(user["id"], user.get("jti", ""), "fulfill_reset", {"request_id": request_id})
        return {"success": True, "temp_password": temp_pw}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)

@app.post("/api/admin/reset-requests/{request_id}/reject")
async def admin_reject_reset(request_id: int, user: dict = Depends(require_admin)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database unavailable")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE password_reset_requests SET status = 'rejected' WHERE id = %s AND status = 'pending'", (request_id,))
        conn.commit()
        cur.close()
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


# ========== TRACKING ENDPOINT ==========

@app.post("/api/track")
async def track_action(req: TrackRequest, user: dict = Depends(get_current_user)):
    if req.action == "search" and (not req.details or not req.details.get("query")):
        # Don't log empty/null search queries
        return {"ok": True}
    log_user_activity(user["id"], user.get("jti", ""), req.action, req.details)
    return {"ok": True}


# ========== PROTECTED DATA ROUTES ==========

def strip_sales_data_for_stock_only(data: dict) -> dict:
    """Remove sensitive sales/financial data for stock_only users. They only see products + stock."""
    import copy
    stripped = copy.deepcopy(data)
    stripped.pop("total_units_sold", None)
    stripped.pop("total_orders", None)
    stripped.pop("total_revenue", None)
    stripped.pop("net_revenue", None)
    stripped.pop("profit", None)
    stripped.pop("profit_margin", None)
    stripped.pop("channel_breakdown", None)
    for p in stripped.get("products", []):
        p.pop("units_sold", None)
        p.pop("total_revenue", None)
        p.pop("sales_by_channel", None)
        p.pop("purchase_orders", None)
    return stripped

@app.get("/api/inventory/search")
async def search_inventory(
    q: str = Query("", description="Search query (name, SKU, BL ID)"),
    user: dict = Depends(get_current_user)
):
    """Search the full inventory (all products/variants) regardless of sales.
    Returns products matching the query by name, SKU, or BL product ID.
    """
    full_inv = cache.get("full_inventory", [])
    if not full_inv:
        return {"products": [], "total": 0}

    if not q or len(q) < 2:
        return {"products": [], "total": 0}

    query_normalized = strip_diacritics(q.lower())
    matches = []
    for p in full_inv:
        if (query_normalized in strip_diacritics((p.get('product_name') or '').lower())
            or query_normalized in strip_diacritics((p.get('sku') or '').lower())
            or query_normalized in str(p.get('bl_product_id', ''))):
            matches.append(p)

    # Enrich with cost data
    costs_by_sku, costs_by_base_sku = cache.get("costs", ({}, {}))
    for p in matches:
        cost, match_type = find_cost_for_sku(p.get('sku', ''), costs_by_sku, costs_by_base_sku)
        p['unit_cost'] = round(cost, 2)
        p['cost_match'] = match_type

    is_stock_only = user.get("role") in ("stock_only", "viewer")
    if is_stock_only:
        for p in matches:
            p.pop('unit_cost', None)
            p.pop('cost_match', None)
            p.pop('total_revenue', None)
            p.pop('units_sold', None)

    return {"products": matches[:50], "total": len(matches)}


@app.get("/api/inventory/full")
async def full_inventory(
    q: str = Query("", description="Search query"),
    category: str = Query("", description="Category filter"),
    stock_filter: str = Query("", description="in_stock, out_of_stock, low_stock"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    user: dict = Depends(get_current_user)
):
    """Full inventory list with search, filters, and pagination.
    Shows ALL products with stock levels - no sales data required.
    """
    full_inv = cache.get("full_inventory", [])
    if not full_inv:
        return {"products": [], "total": 0, "page": 1, "pages": 1,
                "categories": [],
                "summary": {"total": 0, "in_stock": 0, "low_stock": 0, "out_of_stock": 0}}

    filtered = list(full_inv)

    # Search filter — supports name, SKU, BL ID, Shopify variant ID, and Shopify URLs
    if q and len(q) >= 1:
        query = q.strip()
        # Extract Shopify variant ID from pasted URLs like ?variant=12345
        variant_match = re.search(r'[?&]variant=(\d+)', query)
        if variant_match:
            vid = variant_match.group(1)
            filtered = [p for p in filtered if str(p.get('shopify_variant_id', '')) == vid]
        else:
            ql = strip_diacritics(query.lower())
            filtered = [p for p in filtered if
                ql in strip_diacritics((p.get('product_name') or '').lower())
                or ql in strip_diacritics((p.get('sku') or '').lower())
                or ql in str(p.get('bl_product_id', ''))
                or ql in str(p.get('shopify_variant_id', ''))]

    # Category filter
    if category:
        filtered = [p for p in filtered if p.get('category') == category]

    # Stock filter
    if stock_filter == 'in_stock':
        filtered = [p for p in filtered if p.get('current_stock', 0) > 0]
    elif stock_filter == 'out_of_stock':
        filtered = [p for p in filtered if p.get('current_stock', 0) <= 0]
    elif stock_filter == 'low_stock':
        filtered = [p for p in filtered if 0 < p.get('current_stock', 0) <= 5]
    elif stock_filter:  # non-empty but not a valid value
        return JSONResponse(status_code=422, content={"error": f"Invalid stock_filter '{stock_filter}'. Use: in_stock, out_of_stock, low_stock"})

    # Sort by stock descending, then by name
    filtered.sort(key=lambda x: (-x.get('current_stock', 0), (x.get('product_name') or '').lower()))

    # Summary stats (from full inventory, not filtered)
    total_all = len(full_inv)
    in_stock = sum(1 for p in full_inv if p.get('current_stock', 0) > 0)
    low_stock_count = sum(1 for p in full_inv if 0 < p.get('current_stock', 0) <= 5)
    out_of_stock = sum(1 for p in full_inv if p.get('current_stock', 0) == 0)

    # Categories for filter dropdown
    categories = sorted(set(p.get('category', 'Other') for p in full_inv))

    # Paginate
    total = len(filtered)
    pages = max(1, (total + per_page - 1) // per_page)
    start = (page - 1) * per_page
    page_items = filtered[start:start + per_page]

    # Strip sensitive data for stock_only users
    is_stock_only = user.get("role") in ("stock_only", "viewer")
    if is_stock_only:
        for p in page_items:
            p.pop('unit_cost', None)
            p.pop('cost_match', None)
            p.pop('total_revenue', None)
            p.pop('units_sold', None)

    return {
        "products": page_items,
        "total": total,
        "page": page,
        "pages": pages,
        "categories": categories,
        "summary": {
            "total": total_all,
            "in_stock": in_stock,
            "low_stock": low_stock_count,
            "out_of_stock": out_of_stock,
        }
    }


@app.get("/api/sales")
async def get_sales(
    date_from: str = Query("", description="Start date YYYY-MM-DD"),
    date_to: str = Query("", description="End date YYYY-MM-DD"),
    user: dict = Depends(get_current_user)
):
    is_stock_only = user.get("role") in ("stock_only", "viewer")

    if cache["data"] is None:
        return JSONResponse(status_code=503, content={"error": "Data not loaded yet. Please wait..."})

    # No date params -> return cached "all time" data (fast path, no regression)
    if not date_from and not date_to:
        result = {
            "last_updated": cache["last_updated"],
            "next_refresh": cache["next_refresh"],
            "is_refreshing": cache["is_refreshing"],
            **cache["data"]
        }
        return strip_sales_data_for_stock_only(result) if is_stock_only else result

    # With date params -> use ALL-status orders (not just Wysłane)
    # so "Yesterday" shows all 12 orders, not just 2 shipped ones
    raw_orders = cache.get("raw_orders") or cache.get("all_recent_orders")
    inventory = cache.get("inventory") or {}
    po_items_map = cache.get("po_items_by_bl_id") or {}
    order_sources = cache.get("order_sources", {})

    if raw_orders is None:
        return JSONResponse(status_code=503, content={"error": "Raw order data not available. Please wait for refresh."})

    # Parse date boundaries as Polish timezone (user is in Poland)
    # "Today" = midnight-to-midnight in Europe/Warsaw, not UTC
    try:
        if date_from:
            dt_from = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=POLISH_TZ)
            ts_from = int(dt_from.timestamp())
        else:
            ts_from = 0
        if date_to:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=POLISH_TZ) + timedelta(days=1)
            ts_to = int(dt_to.timestamp())
        else:
            ts_to = int(time.time()) + 86400
    except ValueError:
        return JSONResponse(status_code=400, content={"error": "Invalid date format. Use YYYY-MM-DD."})

    if ts_from > ts_to:
        return JSONResponse(status_code=400, content={"error": "date_from must be before date_to"})

    # Filter orders by date_add (order creation time) — NOT date_confirmed
    # date_add is the actual order date; date_confirmed can be days later
    # Exclude cancelled orders — they should never count toward financials
    filtered_orders = []
    for order in raw_orders:
        order_ts = order.get("date_add", 0)
        order_status = order.get("order_status_id") if order.get("order_status_id") is not None else order.get("status_id")
        if order_status not in FINANCIAL_STATUS_IDS:
            continue
        if isinstance(order_ts, (int, float)) and ts_from <= order_ts < ts_to:
            filtered_orders.append(order)

    response_data = build_response(filtered_orders, inventory, po_items_map, order_sources)

    result = {
        "last_updated": cache["last_updated"],
        "next_refresh": cache["next_refresh"],
        "is_refreshing": cache["is_refreshing"],
        "date_from": date_from,
        "date_to": date_to,
        **response_data
    }
    return strip_sales_data_for_stock_only(result) if is_stock_only else result


@app.get("/api/purchase-orders")
async def get_purchase_orders(user: dict = Depends(get_current_user)):
    """Return the list of purchase orders"""
    return {
        "purchase_orders": cache.get("purchase_orders", [])
    }


@app.post("/api/refresh")
async def force_refresh(background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    if cache["is_refreshing"]:
        return {"status": "already_refreshing", "message": "Refresh already in progress"}

    log_activity("manual_refresh_requested")
    background_tasks.add_task(refresh_data)
    return {"status": "started", "message": "Refresh started in background"}


@app.get("/api/activity")
async def get_activity(user: dict = Depends(get_current_user)):
    """Get recent activity log"""
    conn = get_db_connection()
    if not conn:
        return {"error": "Database not connected", "activities": []}
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT timestamp, action, total_orders, total_variants, total_units_sold
            FROM activity_log ORDER BY timestamp DESC LIMIT 50
        """)
        rows = cur.fetchall()
        cur.close()
        return {
            "activities": [
                {
                    "timestamp": r[0].isoformat() if r[0] else None,
                    "action": r[1],
                    "total_orders": r[2],
                    "total_variants": r[3],
                    "total_units_sold": r[4]
                } for r in rows
            ]
        }
    except Exception as e:
        return {"error": str(e), "activities": []}
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


@app.get("/api/debug/costs")
async def debug_costs(user: dict = Depends(require_admin)):
    """Debug: show cost and PO loading status"""
    costs = cache.get("costs", ({}, {}))
    po_items = cache.get("po_items_by_bl_id", {})
    pos = cache.get("purchase_orders", [])
    return {
        "google_credentials_set": bool(GOOGLE_CREDENTIALS_JSON),
        "import_sheet_id": IMPORT_SHEET_ID,
        "costs_by_sku_count": len(costs[0]) if costs else 0,
        "costs_by_base_sku_count": len(costs[1]) if costs else 0,
        "po_items_by_bl_id_count": len(po_items),
        "purchase_orders_count": len(pos),
        "purchase_orders": [
            {"name": po.get("document_number"), "items": po.get("items_count"), "bl_ids": po.get("product_ids", [])[:5]}
            for po in pos[:10]
        ],
        "sample_costs": dict(list((costs[0] if costs else {}).items())[:10]),
        "sample_bl_ids": list(po_items.keys())[:20],
    }


@app.get("/api/debug/orders")
async def debug_orders(user: dict = Depends(require_admin)):
    """Debug: show timestamp fields for recent orders to diagnose date filtering"""
    # Use all-status orders for debug (shows the full picture)
    all_recent = cache.get("all_recent_orders")
    raw_orders = cache.get("raw_orders")
    debug_source = all_recent or raw_orders
    if not debug_source:
        return {"error": "No orders cached"}

    # Get the 20 most recent orders by date_add
    sorted_orders = sorted(debug_source, key=lambda o: o.get("date_add", 0), reverse=True)[:20]

    debug_list = []
    for o in sorted_orders:
        date_add = o.get("date_add", 0)
        date_confirmed = o.get("date_confirmed", 0)
        date_in_status = o.get("date_in_status", 0)

        debug_list.append({
            "order_id": o.get("order_id"),
            "date_add": date_add,
            "date_add_human": datetime.fromtimestamp(date_add, tz=POLISH_TZ).strftime("%Y-%m-%d %H:%M:%S") if date_add else None,
            "date_confirmed": date_confirmed,
            "date_confirmed_human": datetime.fromtimestamp(date_confirmed, tz=POLISH_TZ).strftime("%Y-%m-%d %H:%M:%S") if date_confirmed else None,
            "date_in_status": date_in_status,
            "date_in_status_human": datetime.fromtimestamp(date_in_status, tz=POLISH_TZ).strftime("%Y-%m-%d %H:%M:%S") if date_in_status else None,
            "order_source": o.get("order_source"),
            "order_source_id": o.get("order_source_id"),
            "products_count": len(o.get("products", [])),
            "products_summary": [
                {
                    "name": p.get("name", "")[:50],
                    "sku": p.get("sku", ""),
                    "qty": p.get("quantity"),
                    "price_brutto": p.get("price_brutto"),
                }
                for p in o.get("products", [])[:3]
            ]
        })

    return {
        "total_cached_financial_orders": len(raw_orders) if raw_orders else 0,
        "total_cached_all_status_orders": len(all_recent) if all_recent else 0,
        "source_names": cache.get("order_sources", {}),
        "recent_orders": debug_list
    }


@app.get("/api/debug/orders-db")
async def debug_orders_db(user: dict = Depends(require_admin)):
    """Debug: show order database stats — how many orders stored vs API"""
    conn = get_db_connection()
    if not conn:
        return {"error": "Database not connected"}
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM bl_orders")
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM bl_orders WHERE status_id = ANY(%s)", (list(FINANCIAL_STATUS_IDS),))
        financial = cur.fetchone()[0]
        cur.execute("SELECT MIN(date_add), MAX(date_add) FROM bl_orders")
        row = cur.fetchone()
        min_date = datetime.fromtimestamp(row[0], tz=POLISH_TZ).strftime("%Y-%m-%d") if row[0] else None
        max_date = datetime.fromtimestamp(row[1], tz=POLISH_TZ).strftime("%Y-%m-%d") if row[1] else None
        cur.execute("""
            SELECT order_source, COUNT(*) as cnt
            FROM bl_orders GROUP BY order_source ORDER BY cnt DESC LIMIT 10
        """)
        sources = [{"source": r[0], "count": r[1]} for r in cur.fetchall()]
        cur.close()

        api_financial = len(cache.get("raw_orders", []) or [])
        return {
            "db_total_orders": total,
            "db_financial_orders": financial,
            "db_oldest_order": min_date,
            "db_newest_order": max_date,
            "db_sources": sources,
            "api_financial_cached": api_financial,
            "extra_from_db": max(0, len(cache.get("raw_orders", []) or []) - len(cache.get("all_recent_orders", []) or [])),
        }
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        put_db_connection(conn)


@app.get("/api/download")
async def download_excel(
    category: str = Query("", description="Filter by category"),
    po: str = Query("", description="Filter by purchase order ID"),
    search: str = Query("", description="Search text or Shopify URL"),
    date_from: str = Query("", description="Start date YYYY-MM-DD"),
    date_to: str = Query("", description="End date YYYY-MM-DD"),
    user: dict = Depends(get_current_user)
):
    if cache["data"] is None:
        return JSONResponse(status_code=503, content={"error": "Data not loaded yet"})

    # Determine source products based on date filter
    if date_from or date_to:
        # Use ALL-status orders for date-filtered views (same as /api/sales)
        raw_orders = cache.get("raw_orders") or cache.get("all_recent_orders")
        inventory = cache.get("inventory") or {}
        po_items_map = cache.get("po_items_by_bl_id") or {}
        order_sources = cache.get("order_sources", {})

        if raw_orders is None:
            return JSONResponse(status_code=503, content={"error": "Raw order data not available"})

        try:
            if date_from:
                dt_from = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=POLISH_TZ)
                ts_from = int(dt_from.timestamp())
            else:
                ts_from = 0
            if date_to:
                dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=POLISH_TZ) + timedelta(days=1)
                ts_to = int(dt_to.timestamp())
            else:
                ts_to = int(time.time()) + 86400
        except ValueError:
            return JSONResponse(status_code=400, content={"error": "Invalid date format. Use YYYY-MM-DD."})

        filtered_orders = [
            o for o in raw_orders
            if ts_from <= o.get("date_add", 0) < ts_to
            and (o.get("order_status_id") if o.get("order_status_id") is not None else o.get("status_id")) in FINANCIAL_STATUS_IDS
        ]
        source_data = build_response(filtered_orders, inventory, po_items_map, order_sources)
    else:
        source_data = cache["data"]

    # Apply category/PO/search filters on top
    products = filter_products(source_data["products"], category, po, search)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:G1')
    ws['A1'] = "BASELINKER SALES REPORT - SHIPPED & PACKED ORDERS"
    ws['A1'].font = Font(bold=True, size=14)

    # Show active filters in subtitle
    filter_parts = []
    if date_from or date_to:
        date_label = f"{date_from or 'start'} to {date_to or 'now'}"
        filter_parts.append(f"Period: {date_label}")
    if category:
        filter_parts.append(f"Category: {category}")
    if po:
        po_data = next((p for p in cache.get('purchase_orders', []) if str(p.get('po_id')) == po), None)
        if po_data:
            filter_parts.append(f"PO: {po_data.get('document_number', po)}")
    if search:
        filter_parts.append(f"Search: {search}")
    filter_text = f" | Filters: {', '.join(filter_parts)}" if filter_parts else ""

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}{filter_text}"

    headers = [
        ("Product Name", 50), ("SKU", 25), ("Category", 18),
        ("Units Sold", 12), ("Revenue (PLN)", 16),
        ("Current Stock", 14), ("Image URL", 40)
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, product in enumerate(products, 5):
        ws.cell(row=row_idx, column=1, value=product['product_name'][:80]).border = thin_border
        ws.cell(row=row_idx, column=2, value=product['sku']).border = thin_border
        ws.cell(row=row_idx, column=3, value=product.get('category', 'Other')).border = thin_border
        units_cell = ws.cell(row=row_idx, column=4, value=product['units_sold'])
        units_cell.border = thin_border
        units_cell.alignment = Alignment(horizontal='center')
        rev_cell = ws.cell(row=row_idx, column=5, value=round(product.get('total_revenue', 0), 2))
        rev_cell.border = thin_border
        rev_cell.alignment = Alignment(horizontal='right')
        rev_cell.number_format = '#,##0.00'
        stock_cell = ws.cell(row=row_idx, column=6, value=product['current_stock'])
        stock_cell.border = thin_border
        stock_cell.alignment = Alignment(horizontal='center')
        if product['current_stock'] == 0:
            stock_cell.font = Font(bold=True, color="FF0000")
        elif product['current_stock'] < 5:
            stock_cell.font = Font(color="FF6600")
        ws.cell(row=row_idx, column=7, value=product.get('image_url', '') or '').border = thin_border

    summary_row = len(products) + 5
    ws.cell(row=summary_row, column=1, value=f"TOTAL: {len(products)} variants").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=sum(p['units_sold'] for p in products)).font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value=round(sum(p.get('total_revenue', 0) for p in products), 2)).font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"baselinker_sales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
